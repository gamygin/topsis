import os
import openpyxl
import skcriteria as skc
from skcriteria.preprocessing import invert_objectives, scalers
from skcriteria.madm import similarity
from skcriteria.pipeline import mkpipe
import datetime
from tkinter import *

from util import *

def select_file():
    global wb, file_name
    filetypes = [
        ("excel files", "*.xlsx")
    ]

    filenames = filedialog.askopenfilenames(
        title='Open files',
        initialdir='/',
        filetypes=filetypes)

    print(filenames[0])
    window.destroy()
    result = get_file_direct(filenames[0], error_label)
    if result == "error":
        return 
    else:
        wb = result[0]
        file_name = result[1]
        return
    


window = Tk()
window.geometry("450x150")
window.title("Ranking")

error_label = Label(window, text="", font= ('Arial 10'))

submit = Button(window, text="Select file to analyze", height=2, width=20, command= select_file)

signature = Label(window, text="Made by: Nikita Gamygin", font= ('Arial 10'))

error_label.pack()

submit.pack(pady= 20)

signature.pack(side=BOTTOM)

window.mainloop()


if __name__ == "__main__":
    sheet = wb.active
    #getting ready to get info except matrix
    criteria = []
    weights = []
    objectives = []
    objectives_str = []
    alternatives = []
    column = 2
    row = 4
    rank_column = 0
    while True:
        #getting all the values except matrix

        #criteria
        criteriaValue = sheet.cell(row= 1, column= column).value
        criteria.append(criteriaValue)

        #weights
        weightValue = sheet.cell(row= 2, column= column).value
        try:
            weights.append(float(weightValue))
        except:
            print('Incorrect weights',"\n")
            break
        
        # objectives (min, max)
        objectiveValue = sheet.cell(row= 3, column= column).value
        if objectiveValue == 'min':
            objectives_str.append("min")
            objectives.append(min)
        elif objectiveValue == 'max':
            objectives_str.append("max")
            objectives.append(max)

        #checking next cell
        column += 1
        #this will be used if the program has already been used on the file
        current_cell = sheet.cell(row= 1, column= column).value
        if current_cell == None:
            if not current_cell or not sheet.cell(row= 2, column= column).value or not sheet.cell(row= 3, column= column).value:
                print(f"criteria: {criteria}, weights: {weights}, objectives: {objectives}")
                #getting compared object names (alternatives)
                while True:
                    alternativeValue = sheet.cell(row= row, column= 1).value
                    alternatives.append(alternativeValue)
                    row += 1
                    if not sheet.cell(row= row, column= 1).value:
                        print(f'alternatives (compared objects): {alternatives}')
                        break
                    final_alternative = row
                rank_column = column
                break

    #getting matrix
    matrix = create_matrix(sheet, row, column)

    # evaluating values, putting results them in the file
    try:
        #evaluating values
        dm = skc.mkdm(matrix, objectives, weights, alternatives, criteria)
        print(dm,"\n")

        pipe = mkpipe(
        invert_objectives.NegateMinimize(),
        scalers.VectorScaler(target="matrix"),  # this scaler transform the matrix
        scalers.SumScaler(target="weights"),  # and this transform the weights
        similarity.TOPSIS(),
        )
        print(pipe,"\n")

        rank = pipe.evaluate(dm)
        print(f"rank: {rank} \n")

        rankings = list(rank.values)
        print(f"rankings: {rankings} \n")

        #putting results header
        if not sheet.cell(row= final_alternative + 3, column= 1).value:
            sheet.cell(row= final_alternative + 3, column= 1).value = "Results:"
        last_row = final_alternative + 4

        #getting to the open space
        while True:
            if sheet.cell(row= last_row, column= 1).value:
                last_row += 7
            else:
                break
        #diplaying timestamp
        sheet.cell(row= last_row, column= 1).value = str(datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
        wb.worksheets[0].column_dimensions["A"].width = 20
        last_row += 1

        #displaying data on time of analysis
        display_snapshot(sheet, last_row, alternatives, criteria, weights, objectives_str, matrix)
        
        #displaying rankings
        rank_column = len(criteria) + 2
        display_rankings(sheet, rank_column, last_row, rankings)

        #getting column name in letters
        index = column_index_to_letters(rank_column)

        #adjusting the width of the column with the results
        wb.worksheets[0].column_dimensions[index].width = 15
        #saving file
        wb.save(file_name)

        win = Tk()
        win.geometry("450x150")
        win.title("Success")

        label = Label(win,text="Your file has been successfully analyzed and altered.", font= ('Arial 13'))
        label.pack(pady=20)

        close = Label(win, text= "Please close this window after your finish reading.", font=('Arial 10'))
        close.pack(pady=20)

        win.mainloop()
    except Exception as error:
        print("Processes stopped: Your file doesn't oblige by the structure given in README.md. Review for any possible holes or incorrect value types.")
        print(f"Error: {error}")
        win = Tk()
        win.geometry("450x150")
        win.title("Success")

        label = Label(win,text="Processes stopped:\nYour file doesn't oblige by the structure given in README.md.\nReview for any possible holes or incorrect value types.", font= ('Arial 13'))
        label.pack(pady=20)

        close = Label(win, text= "Please close this window after your finish reading.", font=('Arial 10'))
        close.pack(pady=20)

        win.mainloop()

    
