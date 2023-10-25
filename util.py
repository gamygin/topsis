import os
import openpyxl
import skcriteria as skc
from skcriteria.preprocessing import invert_objectives, scalers
from skcriteria.madm import similarity
from skcriteria.pipeline import mkpipe
import datetime
from tkinter import filedialog


def get_file_direct(file_path, error_label):
    folder_path = file_path[0:file_path.rfind("/")]
    file_name = file_path[file_path.rfind("/") +1:]
    print(folder_path)
    try:
        os.chdir(folder_path)
        wb = openpyxl.load_workbook(file_name)
        return wb, file_name
    except Exception as error:
        print('An error has occured. Please try again.',"\n")
        print(error)
        error_label.config(text = 'An error has occured. Please try again.')
        return "error"

def create_matrix(sheet, row, column):
    matrix = []
    for r in range(row-4):
        values = []
        for c in range(column-2):
            try:
                value = float(sheet.cell(row=r+4, column=c+2).value)
            except:
                print('Incorrect values',"\n")
                break
            values.append(value)
        matrix.append(values)
    print(matrix,"\n")
    return matrix

def display_snapshot(sheet, last_row, alternatives, criteria, weights, objectives, matrix):
    #display the alternatives
    i=0
    print("alternatives:")
    for alternative in alternatives:
            sheet.cell(row= last_row + 3 + i, column= 1).value = alternative
            print(" ",alternative)
            i+=1

    #display the criteria
    i=0
    print("criteria:")
    for criterion in criteria:
        sheet.cell(row= last_row, column= 2 + i).value = criterion
        print(" ",criterion)
        i+=1

    #display the weights
    i=0
    print("weights:")
    for weight in weights:
        sheet.cell(row= last_row + 1, column= 2 + i).value = weight
        print(" ",weight)
        i+=1

    #display the objectives
    i=0
    print("objectives:")
    for objective in objectives:
        sheet.cell(row= last_row + 2, column= 2 + i).value = objective
        print(" ",objective)
        i+=1

    #display matrix
    i=0
    print("matrix:")
    for row in matrix:
        v=0
        for value in row:
            sheet.cell(row= last_row + 3 + i, column= 2 + v).value = value
            print(" ",value)
            v+=1
        i+=1

def display_rankings(sheet, rank_column, last_row, rankings):
    #inserting rank header with date & time
    sheet.cell(row= last_row, column=rank_column).value = f'Rankings:'
    sheet.cell(row= last_row, column=rank_column).fill = openpyxl.styles.fills.PatternFill(patternType='solid', fgColor=openpyxl.styles.colors.Color(rgb='00FFFF00'))
    #inserting ranks
    for rank in rankings:
        sheet.cell(row= last_row + 3 + rankings.index(rank), column=rank_column).value = rank

def column_index_to_letters(rank_column):
    alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"

    if rank_column <= len(alphabet):
        index = alphabet[rank_column-1]
    else:
        t = rank_column//len(alphabet) 
        for i in range(t):
            index = "Z"*t
        index += alphabet[rank_column-len(alphabet*t)-1]

    return index
    